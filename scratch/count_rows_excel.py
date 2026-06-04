import openpyxl

file_path = "/Users/dieptuhuy/Library/CloudStorage/GoogleDrive-dieptuhuy80@gmail.com/Other computers/My Computer 3/D:/Study/System_Design/tests/Testing_Document.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active
print(f"Total rows in Excel: {ws.max_row}")

# Show the last 10 rows
for r_idx in range(max(1, ws.max_row - 10), ws.max_row + 1):
    row_vals = [cell.value for cell in ws[r_idx]]
    print(f"Row {r_idx}: {row_vals[:3]} ... {row_vals[-2:]}")
