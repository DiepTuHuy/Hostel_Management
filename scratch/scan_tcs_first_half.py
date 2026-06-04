import openpyxl

file_path = "/Users/dieptuhuy/Library/CloudStorage/GoogleDrive-dieptuhuy80@gmail.com/Other computers/My Computer 3/D:/Study/System_Design/tests/Testing_Document.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"{'ID':<6} | {'Function':<30} | {'Expected Result':<60}")
print("-" * 105)
for r_idx in range(2, 62):
    tc_id = ws.cell(r_idx, 1).value or ""
    func = ws.cell(r_idx, 2).value or ""
    expected = ws.cell(r_idx, 5).value or ""
    # print first 60 chars of expected
    expected_short = expected.replace("\n", " ")[:60]
    print(f"{tc_id:<6} | {func:<30} | {expected_short:<60}")
