import openpyxl

file_path = "/Users/dieptuhuy/Library/CloudStorage/GoogleDrive-dieptuhuy80@gmail.com/Other computers/My Computer 3/D:/Study/System_Design/tests/Testing_Document.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=== Scanning Testing_Document.xlsx for Zalo, SMS, OTP ===")
count = 0
for r_idx in range(1, ws.max_row + 1):
    row_vals = [str(cell.value or "") for cell in ws[r_idx]]
    row_text = " | ".join(row_vals)
    if any(term in row_text.lower() for term in ["zalo", "sms", "otp"]):
        print(f"Row {r_idx:03d} | ID: {ws.cell(r_idx, 1).value} | Function: {ws.cell(r_idx, 2).value} | Text snippet: '{row_text[:120]}...'")
        count += 1

print(f"Found {count} rows matching the terms.")
