import openpyxl

file_path = "/Users/dieptuhuy/Library/CloudStorage/GoogleDrive-dieptuhuy80@gmail.com/Other computers/My Computer 3/D:/Study/System_Design/tests/Testing_Document.xlsx"
wb = openpyxl.load_workbook(file_path)

print(f"Sheet names: {wb.sheetnames}")
ws = wb.active
print(f"Active sheet: {ws.title}")

# Print first few rows to see template headers
for r_idx in range(1, 10):
    row_vals = [cell.value for cell in ws[r_idx]]
    if any(row_vals):
        print(f"Row {r_idx}: {row_vals}")
