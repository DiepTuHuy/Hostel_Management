import openpyxl

file_path = "/Users/dieptuhuy/Library/CloudStorage/GoogleDrive-dieptuhuy80@gmail.com/Other computers/My Computer 3/D:/Study/System_Design/tests/Testing_Document.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

categories = {}
for r_idx in range(2, ws.max_row + 1):
    tc_id = ws.cell(r_idx, 1).value
    function_name = ws.cell(r_idx, 2).value
    status = ws.cell(r_idx, 7).value
    
    if function_name:
        categories[function_name] = categories.get(function_name, 0) + 1

print(f"Total test cases: {len(categories)}")
for k, v in sorted(categories.items()):
    print(f" - {k}: {v} test case(s)")
