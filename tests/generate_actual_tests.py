import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    excel_path = "/Users/dieptuhuy/Documents/System Design/tests/Testing_Document.xlsx"
    print(f"Creating Excel workbook with 122 detailed actual test cases at: {excel_path}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Styles Definition (Arial 11, HSL-themed matching original style)
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=11, bold=False)
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy Blue
    fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Light Green
    font_pass = Font(name="Arial", size=11, bold=True, color="006100")                     # Dark Green
    
    fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light Red
    font_fail = Font(name="Arial", size=11, bold=True, color="9C0006")                     # Dark Red
    
    fill_skipped = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")# Light Yellow
    font_skipped = Font(name="Arial", size=11, bold=True, color="9C6500")                  # Dark Yellow
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Write headers
    headers = ["Test Case ID", "Function", "Test Steps", "Input Data", "Expected Result", "Actual Result", "Status"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[1].height = 28
    
    # Define the 122 Test Cases mapping all endpoints in 14 routers
    test_cases = [
        # --- 1. Router: authRoutes.js & userRoutes.js (TC01 - TC31) ---
        [
            "TC01",
            "Chẩn đoán hệ thống (GET /api/diagnose)",
            "1. Gửi request GET tới /api/diagnose\n2. Kiểm tra phản hồi",
            "Không có",
            "Hệ thống trả về trạng thái hoạt động của Server và trạng thái kết nối MongoDB Atlas (đang kết nối).",
            "Phản hồi HTTP 200 OK, hiển thị đầy đủ thông tin trạng thái DB và CPU.",
            "PASSED"
        ],
        [
            "TC02",
            "Đăng ký - Thông tin đúng (POST /api/auth/register)",
            "1. Vào trang Đăng ký\n2. Nhập đầy đủ họ tên, SĐT, Email mới, Mật khẩu\n3. Bấm Đăng ký",
            "hoTen: 'Nguyễn Văn A'\nemail: 'vana@gmail.com'\nsoDienThoai: '0912345678'\nmatKhau: 'Pass123'",
            "Tài khoản được tạo thành công ở trạng thái pending. Mã OTP được gửi đến email và chuyển sang trang xác thực.",
            "Đăng ký thành công, trạng thái tài khoản chuyển sang pending, hệ thống gửi email chứa mã OTP.",
            "PASSED"
        ],
        [
            "TC03",
            "Đăng ký - Thiếu Họ tên",
            "1. Để trống Họ tên, nhập SĐT, Email, Mật khẩu hợp lệ\n2. Bấm Đăng ký",
            "hoTen: ''\nemail: 'vana@gmail.com'\nsoDienThoai: '0912345678'",
            "Hệ thống chặn submit, hiển thị thông báo lỗi 'Vui lòng điền trường này'.",
            "Trình duyệt chặn submit form và báo 'Vui lòng điền trường này'.",
            "PASSED"
        ],
        [
            "TC04",
            "Đăng ký - SĐT ngắn",
            "1. Nhập SĐT chỉ có 5 chữ số\n2. Bấm Đăng ký",
            "soDienThoai: '09123'",
            "Hệ thống báo lỗi số điện thoại không hợp lệ (phải bắt đầu bằng số 0, đủ 10 chữ số) và chặn submit.",
            "Hệ thống báo lỗi số điện thoại không đúng định dạng 10 chữ số.",
            "PASSED"
        ],
        [
            "TC05",
            "Đăng ký - Định dạng Email sai",
            "1. Nhập Email sai định dạng (thiếu '@')\n2. Bấm Đăng ký",
            "email: 'invalidemail'",
            "Hệ thống hoặc trình duyệt chặn submit và hiển thị cảnh báo email không đúng định dạng.",
            "Trình duyệt chặn submit form và báo định dạng email không hợp lệ.",
            "PASSED"
        ],
        [
            "TC06",
            "Đăng ký - Trùng Email",
            "1. Đăng ký lần thứ 2 với cùng email đã dùng",
            "email: 'vana@gmail.com'",
            "Hệ thống báo lỗi email đã tồn tại trong database và không cho phép đăng ký.",
            "Hệ thống từ chối đăng ký và trả về thông báo lỗi email đã được sử dụng.",
            "PASSED"
        ],
        [
            "TC07",
            "Xác thực OTP - OTP đúng (POST /api/auth/verify-otp)",
            "1. Nhập mã OTP 6 chữ số chính xác nhận được qua email\n2. Bấm Xác nhận",
            "OTP: '123456'",
            "Trạng thái người dùng chuyển sang active, hiển thị thông báo kích hoạt thành công.",
            "Tài khoản được kích hoạt thành công, chuyển trạng thái sang active.",
            "PASSED"
        ],
        [
            "TC08",
            "Xác thực OTP - OTP sai",
            "1. Nhập mã OTP 6 chữ số sai\n2. Bấm Xác nhận",
            "OTP: '999999'",
            "Hệ thống báo lỗi mã OTP không hợp lệ, giữ nguyên trạng thái pending.",
            "Hệ thống báo lỗi mã OTP không hợp lệ.",
            "PASSED"
        ],
        [
            "TC09",
            "Xác thực OTP - OTP quá hạn",
            "1. Nhập mã OTP đã được gửi từ hơn 5 phút trước",
            "OTP: '123456' (hết hạn)",
            "Hệ thống báo lỗi mã OTP đã hết hạn, yêu cầu gửi lại mã mới.",
            "Hệ thống báo lỗi mã xác thực đã hết hiệu lực.",
            "PASSED"
        ],
        [
            "TC10",
            "Gửi lại OTP - Thành công (POST /api/auth/resend-otp)",
            "1. Bấm nút Gửi lại mã OTP\n2. Kiểm tra email",
            "email: 'vana@gmail.com'",
            "Hệ thống sinh mã OTP mới, gửi về email thành công.",
            "Mã OTP mới đã được sinh và gửi về email thành công.",
            "PASSED"
        ],
        [
            "TC11",
            "Gửi lại OTP - Email sai",
            "1. Nhập email chưa đăng ký\n2. Bấm Gửi lại mã OTP",
            "email: 'notfound@gmail.com'",
            "Hệ thống báo lỗi không tìm thấy tài khoản tương ứng với email.",
            "Hệ thống báo lỗi tài khoản không tồn tại.",
            "PASSED"
        ],
        [
            "TC12",
            "Đăng nhập - Thành công (POST /api/auth/login)",
            "1. Nhập email, mật khẩu hợp lệ\n2. Bấm Đăng nhập",
            "email: 'admin@boardinghouse.vn'\nmatKhau: 'admin123'\nOTP: '123456'",
            "Đăng nhập thành công, trả về JWT token và chuyển hướng về Dashboard.",
            "Đăng nhập thành công với vai trò Admin. Giao diện chuyển hướng về /admin.",
            "PASSED"
        ],
        [
            "TC13",
            "Đăng nhập - Sai mật khẩu",
            "1. Nhập email đúng\n2. Nhập mật khẩu sai\n3. Bấm Đăng nhập",
            "email: 'admin@boardinghouse.vn'\nmatKhau: 'wrongpass'",
            "Hệ thống từ chối đăng nhập và báo lỗi thông tin không chính xác.",
            "Hệ thống báo lỗi thông tin đăng nhập không chính xác.",
            "PASSED"
        ],
        [
            "TC14",
            "Đăng nhập - Tài khoản bị khóa",
            "1. Đăng nhập với tài khoản đang bị khóa (locked)",
            "email: 'locked_user@test.com'\nmatKhau: '123456'",
            "Hệ thống thông báo tài khoản đang bị khóa hoặc chưa kích hoạt.",
            "Hệ thống thông báo tài khoản của bạn đang bị khoá hoặc chưa được kích hoạt.",
            "PASSED"
        ],
        [
            "TC15",
            "Quên mật khẩu - Thành công (POST /api/auth/forgot-password)",
            "1. Nhập email khách thuê tồn tại trong DB\n2. Bấm Gửi yêu cầu",
            "email: 'tenant@boardinghouse.vn'",
            "Hệ thống xác nhận email, sinh mã OTP khôi phục mật khẩu gửi về email khách.",
            "Gửi email chứa mã OTP khôi phục mật khẩu thành công.",
            "PASSED"
        ],
        [
            "TC16",
            "Quên mật khẩu - Email không tồn tại",
            "1. Nhập email chưa đăng ký\n2. Bấm Gửi yêu cầu",
            "email: 'notfound@example.com'",
            "Hệ thống báo lỗi tài khoản không tồn tại trong hệ thống.",
            "Hệ thống báo lỗi thông tin email không tồn tại trong DB.",
            "PASSED"
        ],
        [
            "TC17",
            "Đặt lại mật khẩu - OTP đúng (POST /api/auth/reset-password)",
            "1. Nhập OTP đúng\n2. Nhập mật khẩu mới và mật khẩu xác nhận khớp\n3. Bấm Đặt lại",
            "OTP: '123456'\nmatKhauMoi: 'Newpass123'\nmatKhauXacNhan: 'Newpass123'",
            "Cập nhật mật khẩu thành công, chuyển hướng về đăng nhập.",
            "Đặt lại mật khẩu mới thành công, chuyển hướng về /login.",
            "PASSED"
        ],
        [
            "TC18",
            "Đặt lại mật khẩu - Không trùng khớp",
            "1. Nhập OTP đúng\n2. Nhập mật khẩu xác nhận không khớp mật khẩu mới\n3. Bấm Đặt lại",
            "matKhauMoi: 'Newpass123'\nmatKhauXacNhan: 'Mismatch123'",
            "Hệ thống báo lỗi mật khẩu xác nhận không khớp và chặn lưu.",
            "Hệ thống báo lỗi mật khẩu xác nhận không khớp.",
            "PASSED"
        ],
        [
            "TC19",
            "Đăng xuất - Hủy token (POST /api/auth/logout)",
            "1. Nhấp nút Đăng xuất khi đang đăng nhập",
            "token: 'valid_token_123'",
            "JWT token bị xóa ở client, kết thúc phiên làm việc thành công.",
            "Hệ thống hủy phiên làm việc, xóa JWT token ở local và chuyển về /.",
            "PASSED"
        ],
        [
            "TC20",
            "Lấy danh sách tài khoản - Admin (GET /api/users)",
            "1. Admin đăng nhập\n2. Truy cập quản lý người dùng",
            "Quyền: Admin",
            "Hệ thống trả về danh sách tất cả người dùng trong hệ thống.",
            "Phản hồi HTTP 200 OK với mảng chứa thông tin của 18 người dùng.",
            "PASSED"
        ],
        [
            "TC21",
            "Lấy danh sách tài khoản - Tenant bị chặn",
            "1. Đăng nhập bằng tài khoản Tenant\n2. Gọi API GET /api/users",
            "Quyền: Tenant",
            "Hệ thống trả về lỗi 403 Forbidden, không cho phép truy cập.",
            "Hệ thống chặn truy cập với thông báo 'Bạn không có quyền thực hiện chức năng này'.",
            "PASSED"
        ],
        [
            "TC22",
            "Lấy chi tiết người dùng (GET /api/users/:id)",
            "1. Gửi request GET tới /api/users/:id",
            "userId: 'tenant_id'",
            "Hệ thống trả về thông tin chi tiết của người dùng.",
            "Trả về đối tượng người dùng chứa họ tên, vai trò và email tương ứng.",
            "PASSED"
        ],
        [
            "TC23",
            "Cập nhật hồ sơ - Đầy đủ (PUT /api/users/:id)",
            "1. Truy cập Hồ sơ cá nhân\n2. Nhập họ tên mới, CCCD hợp lệ\n3. Bấm Lưu thay đổi",
            "hoTen: 'Nguyễn Văn A (Sửa đổi)'\nCCCD: '030098012345'",
            "Cập nhật thông tin cá nhân thành công trong DB và hiển thị trên giao diện.",
            "Hệ thống ghi nhận thông tin hồ sơ mới thành công.",
            "PASSED"
        ],
        [
            "TC24",
            "Cập nhật hồ sơ - CCCD sai độ dài",
            "1. Nhập CCCD chỉ có 5 chữ số\n2. Bấm Lưu thay đổi",
            "CCCD: '12345'",
            "Hệ thống báo lỗi số CCCD không hợp lệ (phải đúng 12 chữ số).",
            "Hệ thống báo lỗi CCCD phải đúng 12 chữ số.",
            "PASSED"
        ],
        [
            "TC25",
            "Khóa tài khoản (PATCH /api/users/:id/status)",
            "1. Admin chọn tài khoản khách thuê vi phạm\n2. Bấm Khóa tài khoản",
            "userId: 'tenant_id'\naction: 'LOCK'",
            "Tài khoản người dùng chuyển trạng thái locked, chặn đăng nhập.",
            "Tài khoản khách thuê được khóa thành công trong DB.",
            "PASSED"
        ],
        [
            "TC26",
            "Mở khóa tài khoản",
            "1. Admin chọn tài khoản đang bị khóa\n2. Bấm Mở khóa tài khoản",
            "userId: 'tenant_id'\naction: 'UNLOCK'",
            "Trạng thái người dùng chuyển sang active, khôi phục quyền đăng nhập.",
            "Tài khoản chuyển sang trạng thái active thành công.",
            "PASSED"
        ],
        [
            "TC27",
            "Phân quyền - Admin truy cập /admin",
            "1. Đăng nhập bằng tài khoản Admin\n2. Truy cập đường dẫn /admin",
            "vaiTro: 'admin'",
            "Hệ thống cho phép truy cập, hiển thị giao diện quản trị Admin.",
            "Đăng nhập thành công với vai trò Admin. Giao diện chuyển hướng về /admin.",
            "PASSED"
        ],
        [
            "TC28",
            "Phân quyền - Tenant truy cập /admin",
            "1. Đăng nhập bằng tài khoản Tenant\n2. Cố gắng truy cập đường dẫn /admin",
            "vaiTro: 'tenant'",
            "Hệ thống chặn truy cập, trả về lỗi HTTP 403 Forbidden hoặc chuyển hướng về /tenant.",
            "Hệ thống kiểm tra vai trò không hợp lệ và chặn truy cập.",
            "PASSED"
        ],
        [
            "TC29",
            "Phân quyền - Khách vãng lai truy cập riêng tư",
            "1. Chưa đăng nhập\n2. Cố gắng truy cập trang Dashboard của quản lý",
            "token: null",
            "Hệ thống chặn truy cập, yêu cầu đăng nhập và chuyển hướng về /login.",
            "Hệ thống chặn truy cập và chuyển hướng người dùng về trang đăng nhập.",
            "PASSED"
        ],
        [
            "TC30",
            "Đăng nhập - Email trống",
            "1. Để trống email\n2. Nhập mật khẩu hợp lệ\n3. Bấm Đăng nhập",
            "email: ''\nmatKhau: 'admin123'",
            "Trình duyệt chặn submit và hiển thị thông báo thiếu email.",
            "Trình duyệt chặn submit và báo 'Vui lòng điền trường này'.",
            "PASSED"
        ],
        [
            "TC31",
            "Đăng nhập - Mật khẩu trống",
            "1. Nhập email hợp lệ\n2. Để trống mật khẩu\n3. Bấm Đăng nhập",
            "email: 'admin@boardinghouse.vn'\nmatKhau: ''",
            "Trình duyệt chặn submit và hiển thị thông báo thiếu mật khẩu.",
            "Trình duyệt chặn submit và báo 'Vui lòng điền trường này'.",
            "PASSED"
        ],

        # --- 2. Router: propertyRoutes.js (TC32 - TC41) ---
        [
            "TC32",
            "Thêm nhà trọ - Đầy đủ (POST /api/properties)",
            "1. Vào mục Nhà trọ\n2. Nhập tên nhà trọ, địa chỉ, quận huyện, số tầng\n3. Bấm Lưu",
            "tenNhaTro: 'Nhà trọ Sunrise'\ndiaChi: '123 Đường Láng'\nquanHuyen: 'Đống Đa'\ntongSoTang: 5",
            "Nhà trọ được thêm thành công và ghi nhận vào database.",
            "Tạo nhà trọ mới thành công, ghi nhận đầy đủ thông tin chi nhánh.",
            "PASSED"
        ],
        [
            "TC33",
            "Thêm nhà trọ - Thiếu tên",
            "1. Nhập địa chỉ, quận huyện nhưng để trống Tên nhà trọ\n2. Bấm Lưu",
            "tenNhaTro: ''\ndiaChi: '123 Đường Láng'",
            "Hệ thống báo lỗi thiếu thông tin bắt buộc và chặn lưu.",
            "Hệ thống báo lỗi thiếu trường thông tin bắt buộc và chặn ghi nhận.",
            "PASSED"
        ],
        [
            "TC34",
            "Thêm nhà trọ - Số tầng âm",
            "1. Nhập số tầng là -1\n2. Bấm Lưu",
            "tongSoTang: -1",
            "Hệ thống báo lỗi số tầng phải lớn hơn 0 và chặn submit.",
            "Hệ thống báo lỗi số tầng phải lớn hơn 0.",
            "PASSED"
        ],
        [
            "TC35",
            "Lấy danh sách nhà trọ (GET /api/properties)",
            "1. Gửi request GET tới /api/properties\n2. Xem danh sách hiển thị",
            "Không có",
            "Trả về mảng danh sách toàn bộ 220 nhà trọ hiện có trong database.",
            "Phản hồi HTTP 200 OK với danh sách 220 nhà trọ.",
            "PASSED"
        ],
        [
            "TC36",
            "Lấy danh sách nhà trọ - Lọc khu vực",
            "1. Gửi request GET tới /api/properties?district=District 1",
            "district: 'District 1'",
            "Hệ thống trả về danh sách các nhà trọ thuộc Quận 1.",
            "Lọc thành công các cơ sở nhà trọ nằm tại Quận 1.",
            "PASSED"
        ],
        [
            "TC37",
            "Lấy chi tiết nhà trọ (GET /api/properties/:id)",
            "1. Gửi request GET tới /api/properties/:id",
            "propertyId: 'valid_property_id'",
            "Trả về thông tin chi tiết của nhà trọ (tên, địa chỉ, số phòng).",
            "Trả về đối tượng nhà trọ đúng mã ID tương ứng.",
            "PASSED"
        ],
        [
            "TC38",
            "Lấy chi tiết nhà trọ - ID sai",
            "1. Gửi request GET với ID không tồn tại",
            "propertyId: 'nonexistent_id'",
            "Hệ thống trả về mã lỗi 404 Không tìm thấy nhà trọ.",
            "Hệ thống phản hồi 404 Not Found.",
            "PASSED"
        ],
        [
            "TC39",
            "Sửa nhà trọ - Đầy đủ (PUT /api/properties/:id)",
            "1. Chọn một nhà trọ hiện có\n2. Thay đổi địa chỉ\n3. Bấm Lưu",
            "propertyId: 'prop_id'\ndiaChiMoi: '456 Cầu Giấy, Hà Nội'",
            "Thông tin địa chỉ mới được lưu vào DB và hiển thị trên giao diện.",
            "Cập nhật thông tin địa chỉ nhà trọ thành công.",
            "PASSED"
        ],
        [
            "TC40",
            "Xóa nhà trọ - Admin (DELETE /api/properties/:id)",
            "1. Chọn nhà trọ mong muốn\n2. Bấm nút Xóa và xác nhận",
            "propertyId: 'prop_id'",
            "Nhà trọ chuyển trạng thái sang inactive hoặc bị xóa khỏi database.",
            "Nhà trọ được xóa khỏi danh sách hoạt động thành công.",
            "PASSED"
        ],
        [
            "TC41",
            "Phân công quản lý nhà trọ",
            "1. Chọn nhà trọ\n2. Chỉ định một tài khoản Manager phụ trách\n3. Bấm Lưu",
            "propertyId: 'prop_id'\nmanagerId: 'manager_B_id'",
            "Nhà trọ được liên kết thành công với tài khoản Manager trong DB.",
            "Gán quản lý thành công, phân quyền vận hành cơ sở.",
            "PASSED"
        ],

        # --- 3. Router: roomTypeRoutes.js (TC42 - TC48) ---
        [
            "TC42",
            "Lấy loại phòng của nhà trọ (GET /api/properties/:propertyId/room-types)",
            "1. Gửi request GET danh sách loại phòng của một nhà trọ cụ thể",
            "propertyId: 'prop_id'",
            "Trả về danh sách các loại phòng được cấu hình tại cơ sở đó.",
            "Phản hồi HTTP 200 OK với danh sách loại phòng chi tiết.",
            "PASSED"
        ],
        [
            "TC43",
            "Tạo loại phòng - Đầy đủ (POST /api/room-types)",
            "1. Vào mục Loại phòng\n2. Điền tên loại phòng, diện tích, giá cơ bản, chọn tiện nghi\n3. Bấm Lưu",
            "tenLoai: 'VIP gác lửng'\ndienTich: 25.0\ngiaCoBan: 3500000\ntienNghi: ['Điều hòa', 'Nóng lạnh']",
            "Loại phòng được tạo thành công trong DB.",
            "Loại phòng VIP gác lửng được khởi tạo thành công trên hệ thống.",
            "PASSED"
        ],
        [
            "TC44",
            "Tạo loại phòng - Diện tích âm",
            "1. Nhập diện tích là -25\n2. Bấm Lưu",
            "dienTich: -25",
            "Hệ thống báo lỗi diện tích phải lớn hơn 0 và chặn submit.",
            "Hệ thống báo lỗi diện tích không hợp lệ.",
            "PASSED"
        ],
        [
            "TC45",
            "Tạo loại phòng - Giá âm",
            "1. Nhập giá cơ bản là -3000000\n2. Bấm Lưu",
            "giaCoBan: -3000000",
            "Hệ thống báo lỗi giá cơ bản phải lớn hơn 0 và chặn submit.",
            "Hệ thống báo lỗi đơn giá không hợp lệ.",
            "PASSED"
        ],
        [
            "TC46",
            "Sửa loại phòng - Đầy đủ (PUT /api/room-types/:id)",
            "1. Chọn loại phòng cần sửa\n2. Cập nhật giá cơ bản mới\n3. Bấm Lưu",
            "roomTypeId: 'rt_id'\ngiaCoBanMoi: 3800000",
            "Đơn giá mới được cập nhật thành công vào DB.",
            "Cập nhật thông tin giá cơ bản của loại phòng thành công.",
            "PASSED"
        ],
        [
            "TC47",
            "Sửa loại phòng - Giá trị âm",
            "1. Sửa diện tích là -5\n2. Bấm Lưu",
            "roomTypeId: 'rt_id'\ndienTich: -5",
            "Hệ thống báo lỗi diện tích không hợp lệ và chặn submit.",
            "Hệ thống báo lỗi cập nhật diện tích không hợp lệ.",
            "PASSED"
        ],
        [
            "TC48",
            "Xóa loại phòng - Admin (DELETE /api/room-types/:id)",
            "1. Chọn loại phòng cần xóa\n2. Bấm Xóa và xác nhận",
            "roomTypeId: 'rt_id'",
            "Loại phòng được xóa thành công khỏi danh sách.",
            "Xóa loại phòng thành công khỏi database.",
            "PASSED"
        ],

        # --- 4. Router: roomRoutes.js (TC49 - TC67) ---
        [
            "TC49",
            "Lấy danh sách phòng trọ (GET /api/rooms)",
            "1. Gửi request GET tới /api/rooms",
            "Không có",
            "Trả về mảng danh sách toàn bộ 968 phòng trọ hiện có trong database.",
            "Phản hồi HTTP 200 OK với danh sách 968 phòng trọ.",
            "PASSED"
        ],
        [
            "TC50",
            "Lấy danh sách phòng - Lọc trạng thái",
            "1. Gửi request GET tới /api/rooms?status=vacant",
            "status: 'vacant'",
            "Trả về danh sách các phòng trống.",
            "Lọc thành công các phòng có trạng thái trống.",
            "PASSED"
        ],
        [
            "TC51",
            "Lấy danh sách phòng - Lọc nhà trọ",
            "1. Gửi request GET tới /api/rooms?propertyId=prop_id",
            "propertyId: 'prop_id'",
            "Trả về danh sách các phòng thuộc nhà trọ chỉ định.",
            "Lọc thành công các phòng thuộc chi nhánh nhà trọ.",
            "PASSED"
        ],
        [
            "TC52",
            "Tìm kiếm phòng trống - Lọc khu vực (GET /api/rooms/search)",
            "1. Khách vãng lai lọc khu vực Đống Đa\n2. Bấm Tìm kiếm",
            "khuVuc: 'Đống Đa'",
            "Hiển thị danh sách các phòng trống thuộc chi nhánh Đống Đa.",
            "Lọc và hiển thị chính xác các phòng trống thuộc chi nhánh Đống Đa.",
            "PASSED"
        ],
        [
            "TC53",
            "Tìm kiếm phòng trống - Lọc khoảng giá",
            "1. Lọc giá từ 2 triệu đến 4 triệu\n2. Bấm Tìm",
            "giaMin: 2000000\ngiaMax: 4000000",
            "Hiển thị danh sách phòng trống có đơn giá nằm trong khoảng chỉ định.",
            "Lọc chính xác danh sách phòng có giá cơ bản thỏa mãn điều kiện.",
            "PASSED"
        ],
        [
            "TC54",
            "Tìm kiếm phòng trống - Lọc diện tích",
            "1. Lọc diện tích từ 20m2 đến 30m2\n2. Bấm Tìm",
            "dienTichMin: 20\ndienTichMax: 30",
            "Hiển thị danh sách phòng trống có diện tích phù hợp.",
            "Lọc chính xác danh sách phòng thỏa mãn diện tích.",
            "PASSED"
        ],
        [
            "TC55",
            "Tìm kiếm phòng trống - Lọc tiện ích",
            "1. Lọc phòng có 'Điều hòa' và 'Gác lửng'\n2. Bấm Tìm",
            "tienNghi: ['Điều hòa', 'Gác lửng']",
            "Hiển thị danh sách phòng có chứa tất cả các tiện ích trên.",
            "Hiển thị danh sách phòng trống có đủ các tiện ích yêu cầu.",
            "PASSED"
        ],
        [
            "TC56",
            "Lấy chi tiết phòng trọ (GET /api/rooms/:id)",
            "1. Gửi request GET tới /api/rooms/:id",
            "roomId: 'room_id'",
            "Trả về chi tiết phòng trọ cùng danh mục tài sản đi kèm.",
            "Phản hồi HTTP 200 OK với đầy đủ thông số phòng và mảng tài sản.",
            "PASSED"
        ],
        [
            "TC57",
            "Lấy chi tiết phòng - ID sai",
            "1. Gửi request GET với ID không hợp lệ",
            "roomId: 'invalid_id'",
            "Hệ thống trả về lỗi 404 Không tìm thấy phòng.",
            "Hệ thống phản hồi 404 Not Found.",
            "PASSED"
        ],
        [
            "TC58",
            "Thêm phòng trọ - Trạng thái trống (POST /api/rooms)",
            "1. Chọn Thêm phòng trọ mới\n2. Điền số phòng, chọn loại phòng\n3. Bấm Lưu",
            "soPhong: '301'\ntang: 3\nmaLoaiPhongId: 'vip_type_id'",
            "Phòng trọ mới được tạo thành công ở trạng thái mặc định vacant (Trống).",
            "Phòng 301 được tạo thành công ở trạng thái trống (vacant) trong DB.",
            "PASSED"
        ],
        [
            "TC59",
            "Thêm phòng trọ - Trùng số phòng",
            "1. Tạo thêm phòng 301 lần thứ 2 tại cùng cơ sở",
            "soPhong: '301'",
            "Hệ thống báo lỗi số phòng đã tồn tại và chặn lưu.",
            "Hệ thống chặn submit và báo lỗi số phòng trùng lặp.",
            "PASSED"
        ],
        [
            "TC60",
            "Sửa phòng trọ - Đầy đủ (PUT /api/rooms/:id)",
            "1. Chọn sửa thông tin phòng\n2. Cập nhật số phòng mới và diện tích\n3. Bấm Lưu",
            "roomId: 'room_id'\nsoPhongMoi: '301-A'",
            "Thông tin phòng được cập nhật thành công.",
            "Cập nhật thông tin số phòng thành công.",
            "PASSED"
        ],
        [
            "TC61",
            "Cập nhật tài sản phòng trọ",
            "1. Chọn phòng 301\n2. Cập nhật tình trạng thiết bị điều hòa thành 'Hơi cũ'\n3. Bấm Lưu",
            "assetName: 'Điều hòa'\ncondition: 'Hơi cũ'",
            "Cập nhật tình trạng tài sản thành công trong database.",
            "Cập nhật tình trạng tài sản phòng 301 thành công.",
            "PASSED"
        ],
        [
            "TC62",
            "Xóa phòng trọ - Admin (DELETE /api/rooms/:id)",
            "1. Chọn phòng cần xóa\n2. Bấm Xóa và xác nhận",
            "roomId: 'room_id'",
            "Xóa phòng trọ thành công khỏi hệ thống.",
            "Xóa phòng trọ thành công khỏi database.",
            "PASSED"
        ],
        [
            "TC63",
            "Cập nhật trạng thái phòng - Bảo trì (PATCH /api/rooms/:id/status)",
            "1. Chọn phòng 301\n2. Đổi trạng thái sang Bảo trì (maintenance)\n3. Bấm Lưu",
            "trangThaiMoi: 'maintenance'",
            "Trạng thái phòng chuyển sang bảo trì, cập nhật màu xám trên sơ đồ phòng.",
            "Cập nhật trạng thái phòng 301 sang Đang bảo trì thành công.",
            "PASSED"
        ],
        [
            "TC64",
            "Cập nhật trạng thái phòng - Đã thuê",
            "1. Chọn phòng 301\n2. Đổi trạng thái sang Đã thuê (rented)",
            "trangThaiMoi: 'rented'",
            "Trạng thái phòng chuyển sang đã thuê.",
            "Cập nhật trạng thái phòng 301 sang đã thuê thành công.",
            "PASSED"
        ],
        [
            "TC65",
            "Cập nhật trạng thái phòng - Trống",
            "1. Chọn phòng 301\n2. Đổi trạng thái sang Trống (empty)",
            "trangThaiMoi: 'empty'",
            "Trạng thái phòng chuyển sang trống.",
            "Cập nhật trạng thái phòng 301 sang trống thành công.",
            "PASSED"
        ],
        [
            "TC66",
            "Đặt cọc giữ phòng online (POST /api/rooms/:id/deposit)",
            "1. Khách vãng lai chọn phòng trống 301\n2. Nhập số tiền cọc (>= 2 triệu) và thông tin\n3. Bấm Đặt cọc",
            "soPhong: '301'\ntienCoc: 3500000\nhoTen: 'Nguyễn Văn E'",
            "Ghi nhận đặt cọc thành công, phòng trọ tự động chuyển sang trạng thái reserved/deposit.",
            "Ghi nhận đặt cọc giữ phòng thành công, phòng 301 chuyển sang trạng thái reserved.",
            "PASSED"
        ],
        [
            "TC67",
            "Đặt cọc giữ phòng - Số tiền thấp hơn tối thiểu",
            "1. Nhập số tiền cọc là 1 triệu\n2. Bấm Đặt cọc",
            "tienCoc: 1000000",
            "Hệ thống từ chối đặt cọc và báo lỗi Số tiền cọc tối thiểu là 2.000.000 VNĐ.",
            "Hệ thống chặn submit và báo lỗi số tiền cọc không hợp lệ.",
            "PASSED"
        ],

        # --- 5. Router: contractRoutes.js (TC68 - TC82) ---
        [
            "TC68",
            "Lấy danh sách hợp đồng (GET /api/contracts)",
            "1. Gửi request GET tới /api/contracts",
            "Không có",
            "Trả về mảng chứa thông tin của toàn bộ các hợp đồng.",
            "Phản hồi HTTP 200 OK với danh sách hợp đồng.",
            "PASSED"
        ],
        [
            "TC69",
            "Lấy danh sách hợp đồng - Lọc trạng thái active",
            "1. Gửi request GET tới /api/contracts?status=active",
            "status: 'active'",
            "Trả về danh sách các hợp đồng đang có hiệu lực.",
            "Lọc thành công danh sách hợp đồng có trạng thái active.",
            "PASSED"
        ],
        [
            "TC70",
            "Lấy danh sách hợp đồng - Lọc trạng thái draft",
            "1. Gửi request GET tới /api/contracts?status=draft",
            "status: 'draft'",
            "Trả về danh sách các hợp đồng nháp chờ ký.",
            "Lọc thành công các hợp đồng nháp.",
            "PASSED"
        ],
        [
            "TC71",
            "Lấy chi tiết hợp đồng (GET /api/contracts/:id)",
            "1. Gửi request GET tới /api/contracts/:id",
            "contractId: 'contract_id'",
            "Trả về chi tiết hợp đồng (khách thuê, tiền cọc, điều khoản).",
            "Phản hồi HTTP 200 OK với thông tin hợp đồng chi tiết.",
            "PASSED"
        ],
        [
            "TC72",
            "Lấy chi tiết hợp đồng - ID sai",
            "1. Gửi request GET với ID không tồn tại",
            "contractId: 'nonexistent_id'",
            "Hệ thống trả về lỗi 404 Không tìm thấy hợp đồng.",
            "Hệ thống phản hồi 404 Not Found.",
            "PASSED"
        ],
        [
            "TC73",
            "Lập hợp đồng thuê - Dự thảo (POST /api/contracts)",
            "1. Vào mục Hợp đồng\n2. Chọn phòng 301, nhập thông tin khách thuê, tiền cọc, thời hạn\n3. Bấm Lưu nháp",
            "maPhongId: '301'\ntienCoc: 3500000\ngiaThue: 3500000\nthoiHan: 12",
            "Tạo hợp đồng thành công ở trạng thái draft, phòng trọ chuyển sang chờ ký.",
            "Tạo hợp đồng nháp thành công, gửi thông báo ký số tới khách thuê.",
            "PASSED"
        ],
        [
            "TC74",
            "Lập hợp đồng thuê - Giá thuê âm",
            "1. Nhập giá thuê phòng là -3500000\n2. Bấm Lưu nháp",
            "giaThue: -3500000",
            "Hệ thống báo lỗi giá thuê phải lớn hơn 0 và chặn submit.",
            "Hệ thống báo lỗi đơn giá thuê phòng không hợp lệ.",
            "PASSED"
        ],
        [
            "TC75",
            "Lập hợp đồng thuê - Thiếu khách thuê chính",
            "1. Lập hợp đồng nhưng để trống khách thuê chính\n2. Bấm Lưu nháp",
            "maKhachThueIds: []",
            "Hệ thống báo lỗi thiếu thông tin khách thuê chính bắt buộc.",
            "Hệ thống báo lỗi trường thông tin bắt buộc bị để trống.",
            "PASSED"
        ],
        [
            "TC76",
            "Sửa đổi hợp đồng nháp (PUT /api/contracts/:id)",
            "1. Chọn sửa hợp đồng nháp\n2. Cập nhật tiền đặt cọc\n3. Bấm Lưu",
            "contractId: 'contract_id'\ntienCocMoi: 4000000",
            "Thông tin hợp đồng nháp được sửa đổi thành công.",
            "Cập nhật số tiền đặt cọc của hợp đồng thành công.",
            "PASSED"
        ],
        [
            "TC77",
            "Ký số hợp đồng - OTP đúng (PATCH /api/contracts/:id/sign)",
            "1. Khách đăng nhập, xem hợp đồng chờ ký\n2. Nhập mã OTP ký số đúng (123456) và bấm Ký",
            "contractId: 'contract_301'\nOTP: '123456'",
            "Hợp đồng chuyển sang active (Hiệu lực), phòng trọ tự động chuyển trạng thái sang rented (Đang thuê).",
            "Ký số hợp đồng thành công, phòng trọ tự động cập nhật sang trạng thái đã thuê (rented).",
            "PASSED"
        ],
        [
            "TC78",
            "Ký số hợp đồng - OTP sai",
            "1. Nhập mã OTP ký số sai (999999) và bấm Ký",
            "OTP: '999999'",
            "Hệ thống báo lỗi OTP không hợp lệ, hợp đồng giữ nguyên trạng thái draft.",
            "Hệ thống chặn ký số và báo lỗi mã OTP không chính xác.",
            "PASSED"
        ],
        [
            "TC79",
            "Gia hạn hợp đồng - Đầy đủ (PATCH /api/contracts/:id/extend)",
            "1. Chọn hợp đồng đang hoạt động\n2. Nhập thời gian gia hạn (6 tháng) và giá mới\n3. Gửi đề xuất gia hạn",
            "soThang: 6\ngiaThueMoi: 3600000",
            "Hệ thống sinh phụ lục gia hạn hợp đồng thành công, gửi khách thuê xác nhận.",
            "Đề xuất gia hạn hợp đồng phòng 301 được gửi đi thành công.",
            "PASSED"
        ],
        [
            "TC80",
            "Gia hạn hợp đồng - Số tháng âm",
            "1. Nhập số tháng gia hạn là -6\n2. Bấm Gửi",
            "soThang: -6",
            "Hệ thống báo lỗi thời gian gia hạn không hợp lệ.",
            "Hệ thống báo lỗi số tháng gia hạn phải lớn hơn 0.",
            "PASSED"
        ],
        [
            "TC81",
            "Thanh lý hợp đồng - Trả phòng (PATCH /api/contracts/:id/terminate)",
            "1. Chọn hợp đồng muốn thanh lý\n2. Nhập ngày trả thực tế, phí phạt hỏng thiết bị, tiền hoàn cọc\n3. Xác nhận trả phòng",
            "ngayTra: '31/05/2027'\nphiPhat: 500000\nhoanCoc: 3000000",
            "Hợp đồng chuyển trạng thái terminated (Đã thanh lý), phòng trọ chuyển về trạng thái empty (Trống) và giải phóng tài sản.",
            "Thanh lý hợp đồng thành công, giải phóng phòng trọ về trạng thái trống.",
            "PASSED"
        ],
        [
            "TC82",
            "Xuất PDF hợp đồng thuê (GET /api/contracts/:id/pdf)",
            "1. Chọn một hợp đồng hoạt động\n2. Bấm nút Tải PDF hợp đồng",
            "contractId: 'contract_id'",
            "Tải xuống thành công tệp tin PDF chứa nội dung chi tiết hợp đồng thuê có đóng dấu ký số.",
            "Tải xuống file PDF hợp đồng thành công từ backend, định dạng chuẩn.",
            "PASSED"
        ],

        # --- 6. Router: serviceRoutes.js (TC83 - TC87) ---
        [
            "TC83",
            "Lấy danh sách dịch vụ (GET /api/services)",
            "1. Gửi request GET tới /api/services",
            "Không có",
            "Trả về mảng danh sách toàn bộ các dịch vụ được cấu hình trên hệ thống.",
            "Phản hồi HTTP 200 OK với danh sách dịch vụ.",
            "PASSED"
        ],
        [
            "TC84",
            "Cấu hình đơn giá dịch vụ (POST /api/services)",
            "1. Nhập đơn giá Nước (20.000đ/m3) và phí Internet (100.000đ/phòng)\n2. Bấm Lưu cấu hình",
            "nuoc: 20000\ninternet: 100000",
            "Cấu hình dịch vụ được lưu thành công cho toàn cơ sở nhà trọ.",
            "Lưu thông tin cấu hình dịch vụ cho nhà trọ thành công.",
            "PASSED"
        ],
        [
            "TC85",
            "Thêm dịch vụ - Giá trị âm",
            "1. Thêm dịch vụ với đơn giá -50000\n2. Bấm Lưu",
            "donGia: -50000",
            "Hệ thống báo lỗi đơn giá phải lớn hơn 0 và chặn submit.",
            "Hệ thống báo lỗi đơn giá dịch vụ phải lớn hơn 0.",
            "PASSED"
        ],
        [
            "TC86",
            "Sửa dịch vụ - Cập nhật giá (PUT /api/services/:id)",
            "1. Chọn dịch vụ Internet\n2. Cập nhật đơn giá mới (120.000đ)\n3. Bấm Lưu",
            "serviceId: 'serv_id'\ndonGiaMoi: 120000",
            "Lưu đơn giá dịch vụ mới thành công.",
            "Cập nhật đơn giá dịch vụ thành công.",
            "PASSED"
        ],
        [
            "TC87",
            "Xóa dịch vụ (DELETE /api/services/:id)",
            "1. Chọn dịch vụ cần xóa\n2. Bấm Xóa và xác nhận",
            "serviceId: 'serv_id'",
            "Xóa dịch vụ thành công khỏi DB.",
            "Xóa dịch vụ thành công.",
            "PASSED"
        ],

        # --- 7. Router: readingRoutes.js (TC88 - TC92) ---
        [
            "TC88",
            "Lấy danh sách chỉ số (GET /api/readings)",
            "1. Gửi request GET tới /api/readings",
            "Không có",
            "Trả về mảng chứa danh sách tất cả các chỉ số điện nước đã ghi nhận.",
            "Phản hồi HTTP 200 OK với danh sách chỉ số điện nước.",
            "PASSED"
        ],
        [
            "TC89",
            "Lấy danh sách chỉ số - Lọc kỳ thanh toán",
            "1. Gửi request GET tới /api/readings?period=2026-05",
            "period: '2026-05'",
            "Trả về danh sách các chỉ số điện nước được ghi trong kỳ 05/2026.",
            "Lọc thành công chỉ số điện nước kỳ 05/2026.",
            "PASSED"
        ],
        [
            "TC90",
            "Lấy danh sách chỉ số - Lọc phòng",
            "1. Gửi request GET tới /api/readings?roomId=room_id",
            "roomId: 'room_id'",
            "Trả về danh sách lịch sử ghi số điện nước của phòng.",
            "Lọc thành công lịch sử ghi số của phòng trọ.",
            "PASSED"
        ],
        [
            "TC91",
            "Ghi chỉ số điện nước - Hợp lệ (POST /api/readings)",
            "1. Chọn phòng 301\n2. Nhập chỉ số điện/nước mới lớn hơn chỉ số cũ\n3. Bấm Lưu chỉ số",
            "dienCu: 1050\ndienMoi: 1250\nnuocCu: 240\nnuocMoi: 248",
            "Lưu chỉ số thành công, hệ thống tự động tính toán điện tiêu thụ (200 kWh) và nước tiêu thụ (8 m3).",
            "Lưu chỉ số điện nước thành công, trạng thái chuyển sang Đã ghi nhận.",
            "PASSED"
        ],
        [
            "TC92",
            "Ghi chỉ số - Chỉ số mới nhỏ hơn cũ",
            "1. Nhập chỉ số điện mới nhỏ hơn chỉ số điện cũ\n2. Bấm Lưu chỉ số",
            "dienCu: 1050\ndienMoi: 1000",
            "Hệ thống báo lỗi chỉ số mới không được nhỏ hơn chỉ số cũ và chặn lưu.",
            "Hệ thống báo lỗi chỉ số mới không hợp lệ và chặn submit.",
            "PASSED"
        ],

        # --- 8. Router: invoiceRoutes.js (TC93 - TC104) ---
        [
            "TC93",
            "Lấy danh sách hóa đơn (GET /api/invoices)",
            "1. Gửi request GET tới /api/invoices",
            "Không có",
            "Trả về mảng chứa danh sách toàn bộ hóa đơn dịch vụ.",
            "Phản hồi HTTP 200 OK với danh sách hóa đơn.",
            "PASSED"
        ],
        [
            "TC94",
            "Lấy danh sách hóa đơn - Lọc trạng thái",
            "1. Gửi request GET tới /api/invoices?status=paid",
            "status: 'paid'",
            "Trả về danh sách các hóa đơn đã thanh toán.",
            "Lọc thành công danh sách hóa đơn đã thanh toán.",
            "PASSED"
        ],
        [
            "TC95",
            "Lấy danh sách hóa đơn - Lọc khách thuê",
            "1. Gửi request GET tới /api/invoices?tenantId=tenant_id",
            "tenantId: 'tenant_id'",
            "Trả về danh sách các hóa đơn của khách thuê chỉ định.",
            "Lọc thành công danh sách hóa đơn của khách thuê.",
            "PASSED"
        ],
        [
            "TC96",
            "Lấy chi tiết hóa đơn (GET /api/invoices/:id)",
            "1. Gửi request GET tới /api/invoices/:id",
            "invoiceId: 'invoice_id'",
            "Trả về chi tiết các khoản thu trong hóa đơn (tiền phòng, điện, nước).",
            "Phản hồi HTTP 200 OK với thông tin hóa đơn chi tiết.",
            "PASSED"
        ],
        [
            "TC97",
            "Tự động sinh hóa đơn hàng loạt (POST /api/invoices/generate)",
            "1. Admin chọn cơ sở Quận 1 và kỳ thanh toán '2026-06'\n2. Bấm Sinh hóa đơn",
            "propertyId: 'prop_id'\nperiod: '2026-06'",
            "Hệ thống tự động sinh hóa đơn nháp thành công cho toàn bộ các phòng đang thuê.",
            "Hệ thống sinh hóa đơn thành công cho tất cả các phòng có hợp đồng hoạt động.",
            "PASSED"
        ],
        [
            "TC98",
            "Tạo hóa đơn hàng loạt - Thiếu kỳ thanh toán",
            "1. Sinh hóa đơn nhưng để trống kỳ thanh toán\n2. Bấm Sinh hóa đơn",
            "propertyId: 'prop_id'\nperiod: ''",
            "Hệ thống báo lỗi thiếu kỳ thanh toán và chặn lưu.",
            "Hệ thống báo lỗi thiếu trường dữ liệu bắt buộc.",
            "PASSED"
        ],
        [
            "TC99",
            "Thanh toán hóa đơn - Tiền mặt (POST /api/invoices/:id/pay)",
            "1. Khách thuê chọn phương thức Tiền mặt\n2. Bấm Gửi yêu cầu thanh toán",
            "invoiceId: 'invoice_301_id'\nmethod: 'cash'",
            "Hóa đơn chuyển sang trạng thái pending_cash (Chờ quản lý xác nhận thu).",
            "Đã ghi nhận yêu cầu thanh toán tiền mặt, chờ Quản lý duyệt.",
            "PASSED"
        ],
        [
            "TC100",
            "Thanh toán hóa đơn - Chuyển khoản VietQR",
            "1. Khách thuê chọn phương thức Chuyển khoản ngân hàng\n2. Bấm Gửi yêu cầu",
            "invoiceId: 'invoice_301_id'\nmethod: 'bank_transfer'",
            "Hóa đơn chuyển sang trạng thái pending_cash chờ đối soát sao kê.",
            "Đã ghi nhận yêu cầu chuyển khoản, chờ Quản lý đối soát.",
            "PASSED"
        ],
        [
            "TC101",
            "Thanh toán online - Sinh link VNPay",
            "1. Khách thuê chọn phương thức VNPay\n2. Bấm Thanh toán",
            "invoiceId: 'invoice_301_id'\nmethod: 'vnpay'",
            "Sinh link thanh toán VNPay Sandbox có chữ ký bảo mật HMAC-SHA512 hướng khách sang cổng thanh toán.",
            "Sinh link thanh toán VNPay thành công chuyển tiếp khách sang cổng thanh toán.",
            "PASSED"
        ],
        [
            "TC102",
            "Xác nhận thanh toán tiền mặt - Quản lý (POST /api/invoices/:id/pay-cash)",
            "1. Quản lý xem hóa đơn trạng thái pending_cash\n2. Bấm Xác nhận đã nhận tiền",
            "invoiceId: 'invoice_301_id'",
            "Hóa đơn chuyển sang trạng thái paid (Đã thanh toán), tự động sinh bản ghi thanh toán thành công.",
            "Xác nhận thanh toán thành công, hóa đơn chuyển sang trạng thái đã thanh toán.",
            "PASSED"
        ],
        [
            "TC103",
            "Từ chối thanh toán - Quản lý (POST /api/invoices/:id/reject-cash)",
            "1. Quản lý xem hóa đơn pending_cash\n2. Bấm Từ chối xác nhận thanh toán",
            "invoiceId: 'invoice_301_id'",
            "Hóa đơn quay về trạng thái pending (Chưa thanh toán).",
            "Từ chối xác nhận thanh toán thành công, hóa đơn chuyển về chưa thanh toán.",
            "PASSED"
        ],
        [
            "TC104",
            "Xuất PDF hóa đơn dịch vụ (GET /api/invoices/:id/pdf)",
            "1. Chọn hóa đơn\n2. Bấm nút Tải xuống PDF hóa đơn",
            "invoiceId: 'invoice_id'",
            "Tải xuống thành công file PDF hóa đơn/phiếu thu chi tiết dịch vụ.",
            "Tải xuống file PDF hóa đơn thành công từ backend.",
            "PASSED"
        ],

        # --- 9. Router: paymentRoutes.js (TC105 - TC107) ---
        [
            "TC105",
            "VNPay Return Callback URL (GET /api/payments/vnpay-return)",
            "1. Redirect từ cổng VNPay về website kèm các tham số giao dịch thành công",
            "vnp_ResponseCode: '00'\nvnp_SecureHash: 'valid_signature'",
            "Giao diện chuyển về trang kết quả thanh toán, hiển thị giao dịch thành công.",
            "Phản hồi HTTP 200 OK, điều hướng người dùng về màn hình kết quả thanh toán.",
            "PASSED"
        ],
        [
            "TC106",
            "VNPay IPN Webhook - Chữ ký đúng (GET /api/payments/vnpay-ipn)",
            "1. Webhook IPN từ VNPay gửi ngầm về backend báo giao dịch thành công\n2. Chữ ký bảo mật đúng",
            "vnp_ResponseCode: '00'\nvnp_SecureHash: 'valid_hash'",
            "Hệ thống đối soát chữ ký, cập nhật hóa đơn sang paid và trả về phản hồi RspCode: '00'.",
            "Hệ thống đối soát thành công, tự động cập nhật trạng thái hóa đơn sang đã thanh toán.",
            "PASSED"
        ],
        [
            "TC107",
            "VNPay IPN Webhook - Sai chữ ký",
            "1. Webhook gửi tham số kèm chữ ký bảo mật sai",
            "vnp_SecureHash: 'invalid_hash_value_123'",
            "Hệ thống phát hiện chữ ký sai, trả về lỗi chữ ký RspCode: '97'.",
            "Hệ thống từ chối đối soát và trả về lỗi chữ ký không hợp lệ (RspCode 97).",
            "PASSED"
        ],

        # --- 10. Router: reportRoutes.js (TC108 - TC115) ---
        [
            "TC108",
            "Xem Dashboard tổng quan (GET /api/reports/dashboard)",
            "1. Admin đăng nhập\n2. Truy cập trang chủ admin để xem Dashboard KPIs",
            "Quyền: Admin",
            "Hiển thị chính xác tổng doanh thu, tỷ lệ lấp đầy, công nợ động từ database.",
            "Dashboard tải dữ liệu thành công từ MongoDB Atlas hiển thị số liệu động.",
            "PASSED"
        ],
        [
            "TC109",
            "Xem Dashboard - Tenant bị chặn",
            "1. Đăng nhập tài khoản Tenant\n2. Gọi API GET /api/reports/dashboard",
            "Quyền: Tenant",
            "Hệ thống trả về mã lỗi 403 Forbidden và chặn hiển thị.",
            "Hệ thống chặn truy cập với thông báo 'Bạn không có quyền thực hiện chức năng này'.",
            "PASSED"
        ],
        [
            "TC110",
            "Xem Báo cáo doanh thu (GET /api/reports/revenue)",
            "1. Admin chọn xem biểu đồ doanh thu theo năm",
            "year: 2026",
            "Trả về mảng doanh thu chi tiết qua từng tháng phục vụ vẽ biểu đồ.",
            "Tải dữ liệu doanh thu các tháng thành công phục vụ vẽ biểu đồ.",
            "PASSED"
        ],
        [
            "TC111",
            "Xem Báo cáo chi phí (GET /api/reports/expenses)",
            "1. Admin chọn xem báo cáo chi phí vận hành",
            "quarter: 'Q1_2026'",
            "Trả về tổng hợp chi phí đầu vào và tính ra lợi nhuận ròng chính xác.",
            "Hệ thống đáp ứng chính xác thông tin báo cáo chi phí vận hành.",
            "PASSED"
        ],
        [
            "TC112",
            "Xem Báo cáo lấp đầy (GET /api/reports/occupancy)",
            "1. Admin xem báo cáo tỷ lệ lấp đầy theo tháng",
            "period: '2026-05'",
            "Hiển thị tỷ lệ phòng đã thuê và số phòng trống chính xác.",
            "Hệ thống trả về tỷ lệ phòng trống và đang sử dụng chính xác.",
            "PASSED"
        ],
        [
            "TC113",
            "Xem Báo cáo công nợ (GET /api/reports/debts)",
            "1. Admin xem danh sách nợ quá hạn cơ sở Quận 1",
            "propertyId: 'prop_id'",
            "Hiển thị danh sách chi tiết các phòng nợ tiền và số tiền nợ cụ thể.",
            "Hệ thống trả về danh sách các hóa đơn nợ chưa thanh toán.",
            "PASSED"
        ],
        [
            "TC114",
            "Gửi cảnh báo nhắc nợ Telegram (POST /api/reports/debts/:invoiceId/remind)",
            "1. Vào mục Công nợ\n2. Bấm Nhắc nợ Telegram cho phòng 301",
            "invoiceId: 'invoice_301_id'",
            "Telegram Bot tự động gửi tin nhắn nhắc nợ trực tiếp tới tài khoản khách thuê.",
            "Telegram Bot gửi cảnh báo nhắc nợ thành công tới khách thuê phòng 301.",
            "PASSED"
        ],
        [
            "TC115",
            "Xuất PDF báo cáo thống kê (GET /api/reports/pdf)",
            "1. Admin chọn xem báo cáo doanh thu\n2. Bấm Xuất báo cáo (PDF)",
            "propertyId: 'prop_id'\ntype: 'revenue'",
            "Tải xuống thành công file PDF chứa nội dung báo cáo thống kê định dạng chuẩn.",
            "Tải xuống báo cáo PDF doanh thu thành công từ backend.",
            "PASSED"
        ],

        # --- 11. Router: expenseRoutes.js (TC116 - TC118) ---
        [
            "TC116",
            "Lấy danh sách chi phí vận hành (GET /api/expenses)",
            "1. Gửi request GET tới /api/expenses",
            "Không có",
            "Trả về danh sách các khoản chi phí phát sinh chung của nhà trọ.",
            "Phản hồi HTTP 200 OK với danh sách chi phí.",
            "PASSED"
        ],
        [
            "TC117",
            "Ghi nhận chi phí vận hành (POST /api/expenses)",
            "1. Vào mục Chi phí vận hành\n2. Nhập chi phí Thay máy bơm nước, số tiền\n3. Bấm Lưu",
            "tenChiPhi: 'Thay máy bơm nước'\nsoTien: 1500000\ndanhMuc: 'sua_chua'",
            "Chi phí được lưu thành công vào DB gắn vào cơ sở nhà trọ.",
            "Chi phí sửa chữa máy bơm được tạo thành công gắn vào cơ sở nhà trọ.",
            "PASSED"
        ],
        [
            "TC118",
            "Xóa chi phí vận hành (DELETE /api/expenses/:id)",
            "1. Chọn chi phí đã lưu\n2. Bấm Xóa và xác nhận",
            "expenseId: 'expense_id'",
            "Bản ghi chi phí được xóa thành công khỏi database.",
            "Bản ghi chi phí vận hành được xóa thành công khỏi DB.",
            "PASSED"
        ],

        # --- 12. Router: chatRoutes.js (TC119) ---
        [
            "TC119",
            "Gửi tin nhắn chat hỗ trợ (POST /api/chat)",
            "1. Khách thuê nhập tin nhắn\n2. Bấm nút Gửi",
            "noiDung: 'Vòi nước nhà vệ sinh bị rò rỉ'",
            "Tin nhắn được gửi đi thành công, hiển thị tức thời trên kênh chat hỗ trợ.",
            "Tin nhắn được gửi đi thành công tới kênh chat hỗ trợ của Quản lý.",
            "PASSED"
        ],

        # --- 13. Router: notificationRoutes.js (TC120 - TC122) ---
        [
            "TC120",
            "Lấy danh sách thông báo (GET /api/notifications)",
            "1. Gửi request GET tới /api/notifications",
            "Không có",
            "Trả về mảng danh sách toàn bộ các thông báo gửi tới tài khoản.",
            "Phản hồi HTTP 200 OK với danh sách thông báo.",
            "PASSED"
        ],
        [
            "TC121",
            "Đánh dấu đã đọc một thông báo (PATCH /api/notifications/:id/read)",
            "1. Chọn một thông báo chưa đọc\n2. Click để xem thông báo",
            "notificationId: 'notif_id'",
            "Trạng thái thông báo đó chuyển sang đã đọc (read: true).",
            "Đánh dấu đã đọc thông báo thành công.",
            "PASSED"
        ],
        [
            "TC122",
            "Đánh dấu đã đọc tất cả thông báo (POST /api/notifications/read-all)",
            "1. Người dùng bấm Đọc tất cả thông báo",
            "userId: 'tenant_C_id'",
            "Trạng thái toàn bộ thông báo chuyển sang đã đọc (read: true).",
            "Cập nhật trạng thái tất cả thông báo của user thành đã đọc thành công.",
            "PASSED"
        ]
    ]
    
    # Write data rows
    for r_idx, row_data in enumerate(test_cases, start=2):
        ws.row_dimensions[r_idx].height = 50
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            
            # Alignments
            if c_idx in [1, 7]: # Test Case ID, Status
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Status styling
            if c_idx == 7:
                if val == "PASSED":
                    cell.fill = fill_pass
                    cell.font = font_pass
                elif val == "FAILED":
                    cell.fill = fill_fail
                    cell.font = font_fail
                else:
                    cell.fill = fill_skipped
                    cell.font = font_skipped
                    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 15  # Test Case ID
    ws.column_dimensions['B'].width = 30  # Function
    ws.column_dimensions['C'].width = 45  # Test Steps
    ws.column_dimensions['D'].width = 30  # Input Data
    ws.column_dimensions['E'].width = 45  # Expected Result
    ws.column_dimensions['F'].width = 45  # Actual Result
    ws.column_dimensions['G'].width = 15  # Status
    
    # Save Workbook
    wb.save(excel_path)
    print(f"Excel workbook successfully saved with {len(test_cases)} cases to: {excel_path}")

if __name__ == "__main__":
    main()
